import psycopg2
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Conexão com o banco
conn = psycopg2.connect(
    host="localhost",
    database="itributos",
    user="postgres",
    password="postgres",
    port="5432"
)

print("✓ Conectado ao banco de dados")

# Query para buscar TODOS os dados de TLLF 2024 e 2025
query = """
SELECT DISTINCT 
    p.id AS id_contribuinte,
    p.name AS nome_contribuinte,
    p.cpf_cnpj,
    td.year AS ano_referencia,
    td.parcel_number AS numero_parcela,
    td.due_date AS data_vencimento,
    td.value AS valor_parcela,
    td.status,
    td.active_debt_status,
    td.payment_id,
    CASE 
        WHEN td.status = 0 THEN 'Simulado'
        WHEN td.status = 1 THEN 'Aberto'
        WHEN td.status = 2 THEN 'Cancelado'
        WHEN td.status = 3 THEN 'Isento'
        WHEN td.status = 4 THEN 'Dívida Ativa'
        WHEN td.status = 5 THEN 'Pago'
        WHEN td.status = 6 THEN 'Anulado'
        WHEN td.status = 7 THEN 'Excluído'
        WHEN td.status = 8 THEN 'Isenção'
        WHEN td.status = 9 THEN 'Imunidade'
        WHEN td.status = 10 THEN 'Incentivo'
        WHEN td.status = 11 THEN 'Remissão'
        WHEN td.status = 12 THEN 'Suspenso'
        WHEN td.status = 13 THEN 'Parcelado'
        WHEN td.status = 14 THEN 'ISS'
        WHEN td.status = 15 THEN 'Sem Movimento'
        WHEN td.status = 16 THEN 'Supervisionado'
        WHEN td.status = 17 THEN 'Doação em Pagamento'
        WHEN td.status = 18 THEN 'Prescrito'
        WHEN td.status = 19 THEN 'Transferido'
        WHEN td.status = 20 THEN 'Anistiado'
        ELSE 'Desconhecido'
    END AS situacao_pagamento
FROM unico_people p
INNER JOIN taxable_debts td ON td.person_id = p.id
WHERE td.revenue_acronym = 'TLLF'
AND td.year IN (2024, 2025)
ORDER BY td.year DESC, p.name, td.parcel_number
"""

print("✓ Executando consulta...")
df = pd.read_sql_query(query, conn)
conn.close()

print(f"✓ Dados carregados: {len(df)} registros")

# Converter data_vencimento para formato legível
df['data_vencimento'] = pd.to_datetime(df['data_vencimento']).dt.strftime('%d/%m/%Y')

# Criar Workbook
wb = Workbook()
wb.remove(wb.active)

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
pago_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde - Pago
aberto_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarelo - Aberto
cancelado_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Cinza - Cancelado
parcelado_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Azul Claro - Parcelado
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============== SHEET 1: RESUMO GERAL ==============
print("✓ Criando sheet: Resumo Geral")
ws_resumo = wb.create_sheet("Resumo Geral")

# Cabeçalho
ws_resumo.merge_cells('A1:H1')
ws_resumo['A1'] = 'RELATÓRIO COMPLETO TLLF - TODOS OS CONTRIBUINTES'
ws_resumo['A1'].font = Font(size=14, bold=True, color="366092")
ws_resumo['A1'].alignment = Alignment(horizontal='center', vertical='center')

ws_resumo['A3'] = 'Período:'
ws_resumo['B3'] = '2024 e 2025'
ws_resumo['A4'] = 'Data do Relatório:'
ws_resumo['B4'] = datetime.now().strftime('%d/%m/%Y %H:%M')
ws_resumo['A5'] = 'Município:'
ws_resumo['B5'] = 'Curionópolis - PA'
ws_resumo['A6'] = 'Total de Registros:'
ws_resumo['B6'] = len(df)

# Calcular estatísticas
stats_2025 = df[df['ano_referencia'] == 2025]
stats_2024 = df[df['ano_referencia'] == 2024]

resumo_data = []
for ano, dados in [(2025, stats_2025), (2024, stats_2024)]:
    total_contrib = dados['id_contribuinte'].nunique()
    total_parcelas = len(dados)
    valor_total = dados['valor_parcela'].sum()
    
    pagas = dados[dados['status'] == 5]
    qtd_pagas = len(pagas)
    valor_pago = pagas['valor_parcela'].sum()
    perc_pago = (qtd_pagas / total_parcelas * 100) if total_parcelas > 0 else 0
    
    parceladas = dados[dados['status'] == 13]
    qtd_parceladas = len(parceladas)
    
    canceladas = dados[dados['status'] == 2]
    qtd_canceladas = len(canceladas)
    valor_cancelado = canceladas['valor_parcela'].sum()
    
    aberto = dados[(dados['status'] != 5) & (dados['status'] != 2) & (dados['status'] != 13)]
    qtd_aberto = len(aberto)
    valor_aberto = aberto['valor_parcela'].sum()
    
    resumo_data.append([
        ano,
        total_contrib,
        total_parcelas,
        f"R$ {valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        qtd_pagas,
        f"{perc_pago:.1f}%",
        qtd_parceladas,
        qtd_aberto,
        f"R$ {valor_aberto:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        qtd_canceladas
    ])

# Adicionar linha de total
total_contrib_geral = df['id_contribuinte'].nunique()
total_parcelas_geral = len(df)
valor_total_geral = df['valor_parcela'].sum()
pagas_geral = df[df['status'] == 5]
qtd_pagas_geral = len(pagas_geral)
valor_pago_geral = pagas_geral['valor_parcela'].sum()
perc_pago_geral = (qtd_pagas_geral / total_parcelas_geral * 100) if total_parcelas_geral > 0 else 0

parceladas_geral = df[df['status'] == 13]
qtd_parceladas_geral = len(parceladas_geral)

aberto_geral = df[(df['status'] != 5) & (df['status'] != 2) & (df['status'] != 13)]
qtd_aberto_geral = len(aberto_geral)
valor_aberto_geral = aberto_geral['valor_parcela'].sum()

canceladas_geral = df[df['status'] == 2]
qtd_canceladas_geral = len(canceladas_geral)

resumo_data.append([
    'TOTAL',
    total_contrib_geral,
    total_parcelas_geral,
    f"R$ {valor_total_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
    qtd_pagas_geral,
    f"{perc_pago_geral:.1f}%",
    qtd_parceladas_geral,
    qtd_aberto_geral,
    f"R$ {valor_aberto_geral:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
    qtd_canceladas_geral
])

# Headers
headers = ['Ano', 'Contribuintes', 'Parcelas', 'Valor Total', 'Qtd Pagas', '% Pagas', 
           'Qtd Parceladas', 'Qtd Em Aberto', 'Valor Em Aberto', 'Qtd Canceladas']

ws_resumo.append([])
ws_resumo.append([])
row_idx = 9
for col, header in enumerate(headers, 1):
    cell = ws_resumo.cell(row=row_idx, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for row_data in resumo_data:
    row_idx += 1
    for col, value in enumerate(row_data, 1):
        cell = ws_resumo.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.alignment = Alignment(horizontal='center' if col <= 2 else 'right')
        if 'TOTAL' in str(row_data[0]):
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

# Ajustar larguras
widths = [10, 15, 12, 18, 12, 10, 15, 18, 15]
for i, width in enumerate(widths, 1):
    ws_resumo.column_dimensions[get_column_letter(i)].width = width

# ============== SHEET 2: TODOS OS DADOS 2025 ==============
print("✓ Criando sheet: TLLF 2025 Completo")
ws_2025 = wb.create_sheet("TLLF 2025 Completo")

ws_2025['A1'] = 'TLLF 2025 - TODOS OS CONTRIBUINTES'
ws_2025['A1'].font = Font(size=12, bold=True, color="366092")
ws_2025.merge_cells('A1:J1')

df_2025 = df[df['ano_referencia'] == 2025].copy()
df_2025['valor_parcela'] = df_2025['valor_parcela'].apply(
    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
)

headers_detalhe = ['ID', 'Nome Contribuinte', 'CPF/CNPJ', 'Nº Parcela', 
                   'Vencimento', 'Valor', 'Status', 'Situação', 'Payment ID', 'Dívida Ativa']

row_idx = 3
for col, header in enumerate(headers_detalhe, 1):
    cell = ws_2025.cell(row=row_idx, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for _, row_data in df_2025.iterrows():
    row_idx += 1
    dados = [
        row_data['id_contribuinte'],
        row_data['nome_contribuinte'],
        row_data['cpf_cnpj'],
        row_data['numero_parcela'],
        row_data['data_vencimento'],
        row_data['valor_parcela'],
        row_data['status'],
        row_data['situacao_pagamento'],
        row_data['payment_id'] if pd.notna(row_data['payment_id']) else '-',
        row_data['active_debt_status'] if pd.notna(row_data['active_debt_status']) else '-'
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_2025.cell(row=row_idx, column=col, value=value)
        cell.border = border
        
        # Aplicar cores
        if row_data['situacao_pagamento'] == 'Pago':
            cell.fill = pago_fill
        elif row_data['situacao_pagamento'] == 'Cancelado':
            cell.fill = cancelado_fill
        elif row_data['situacao_pagamento'] == 'Parcelado':
            cell.fill = parcelado_fill
        elif row_data['situacao_pagamento'] == 'Aberto':
            cell.fill = aberto_fill
        
        if col in [5]:  # Vencimento
            cell.alignment = Alignment(horizontal='center')
        elif col in [6]:  # Valor
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')

widths_detalhe = [8, 40, 20, 10, 12, 16, 8, 14, 12, 12]
for i, width in enumerate(widths_detalhe, 1):
    ws_2025.column_dimensions[get_column_letter(i)].width = width

# ============== SHEET 3: TODOS OS DADOS 2024 ==============
print("✓ Criando sheet: TLLF 2024 Completo")
ws_2024 = wb.create_sheet("TLLF 2024 Completo")

ws_2024['A1'] = 'TLLF 2024 - TODOS OS CONTRIBUINTES'
ws_2024['A1'].font = Font(size=12, bold=True, color="366092")
ws_2024.merge_cells('A1:J1')

df_2024 = df[df['ano_referencia'] == 2024].copy()
df_2024['valor_parcela'] = df_2024['valor_parcela'].apply(
    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
)

row_idx = 3
for col, header in enumerate(headers_detalhe, 1):
    cell = ws_2024.cell(row=row_idx, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for _, row_data in df_2024.iterrows():
    row_idx += 1
    dados = [
        row_data['id_contribuinte'],
        row_data['nome_contribuinte'],
        row_data['cpf_cnpj'],
        row_data['numero_parcela'],
        row_data['data_vencimento'],
        row_data['valor_parcela'],
        row_data['status'],
        row_data['situacao_pagamento'],
        row_data['payment_id'] if pd.notna(row_data['payment_id']) else '-',
        row_data['active_debt_status'] if pd.notna(row_data['active_debt_status']) else '-'
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_2024.cell(row=row_idx, column=col, value=value)
        cell.border = border
        
        if row_data['situacao_pagamento'] == 'Pago':
            cell.fill = pago_fill
        elif row_data['situacao_pagamento'] == 'Cancelado':
            cell.fill = cancelado_fill
        elif row_data['situacao_pagamento'] == 'Parcelado':
            cell.fill = parcelado_fill
        elif row_data['situacao_pagamento'] == 'Aberto':
            cell.fill = aberto_fill
        
        if col in [5]:
            cell.alignment = Alignment(horizontal='center')
        elif col in [6]:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')

for i, width in enumerate(widths_detalhe, 1):
    ws_2024.column_dimensions[get_column_letter(i)].width = width

# ============== SHEET 4: SOMENTE EM ABERTO ==============
print("✓ Criando sheet: Em Aberto")
ws_aberto = wb.create_sheet("Em Aberto")

ws_aberto['A1'] = 'TLLF EM ABERTO - STATUS DIFERENTE DE PAGO E CANCELADO'
ws_aberto['A1'].font = Font(size=12, bold=True, color="C00000")
ws_aberto.merge_cells('A1:J1')

df_aberto = df[(df['status'] != 5) & (df['status'] != 2)].copy()
df_aberto['valor_parcela'] = df_aberto['valor_parcela'].apply(
    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
)

row_idx = 3
for col, header in enumerate(headers_detalhe, 1):
    cell = ws_aberto.cell(row=row_idx, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for _, row_data in df_aberto.iterrows():
    row_idx += 1
    dados = [
        row_data['id_contribuinte'],
        row_data['nome_contribuinte'],
        row_data['cpf_cnpj'],
        row_data['numero_parcela'],
        row_data['data_vencimento'],
        row_data['valor_parcela'],
        row_data['status'],
        row_data['situacao_pagamento'],
        row_data['payment_id'] if pd.notna(row_data['payment_id']) else '-',
        row_data['active_debt_status'] if pd.notna(row_data['active_debt_status']) else '-'
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_aberto.cell(row=row_idx, column=col, value=value)
        cell.border = border
        
        # Aplicar cores baseado no status
        if row_data['situacao_pagamento'] == 'Aberto':
            cell.fill = aberto_fill
        elif row_data['situacao_pagamento'] == 'Parcelado':
            cell.fill = parcelado_fill
        
        if col in [5]:
            cell.alignment = Alignment(horizontal='center')
        elif col in [6]:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')

for i, width in enumerate(widths_detalhe, 1):
    ws_aberto.column_dimensions[get_column_letter(i)].width = width

# ============== SHEET 5: ANÁLISE POR CONTRIBUINTE ==============
print("✓ Criando sheet: Análise por Contribuinte")
ws_analise = wb.create_sheet("Análise por Contribuinte")

ws_analise['A1'] = 'ANÁLISE CONSOLIDADA POR CONTRIBUINTE'
ws_analise['A1'].font = Font(size=12, bold=True, color="366092")
ws_analise.merge_cells('A1:I1')

# Agrupar por contribuinte
df_agrupado = df.groupby(['id_contribuinte', 'nome_contribuinte', 'cpf_cnpj']).agg({
    'valor_parcela': ['count', 'sum'],
}).reset_index()

df_agrupado.columns = ['id_contribuinte', 'nome_contribuinte', 'cpf_cnpj', 'qtd_parcelas', 'valor_total']

# Contar por status
for status_val, status_nome in [(5, 'qtd_pagas'), (2, 'qtd_canceladas'), (13, 'qtd_parceladas'), (1, 'qtd_status1')]:
    df_status = df[df['status'] == status_val].groupby('id_contribuinte').size().reset_index(name=status_nome)
    df_agrupado = df_agrupado.merge(df_status, on='id_contribuinte', how='left')
    df_agrupado[status_nome] = df_agrupado[status_nome].fillna(0).astype(int)

# Ordenar por valor total
df_agrupado = df_agrupado.sort_values('valor_total', ascending=False)

# Formatar valores
df_agrupado['valor_total'] = df_agrupado['valor_total'].apply(
    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
)

headers_analise = ['ID', 'Nome', 'CPF/CNPJ', 'Total Parcelas', 'Valor Total', 
                   'Qtd Pagas', 'Qtd Canceladas', 'Qtd Parceladas', 'Qtd Status 1']

row_idx = 3
for col, header in enumerate(headers_analise, 1):
    cell = ws_analise.cell(row=row_idx, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = border

for _, row_data in df_agrupado.iterrows():
    row_idx += 1
    dados = [
        row_data['id_contribuinte'],
        row_data['nome_contribuinte'],
        row_data['cpf_cnpj'],
        row_data['qtd_parcelas'],
        row_data['valor_total'],
        row_data['qtd_pagas'],
        row_data['qtd_canceladas'],
        row_data['qtd_parceladas'],
        row_data['qtd_status1']
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_analise.cell(row=row_idx, column=col, value=value)
        cell.border = border
        
        if col in [1, 4]:
            cell.alignment = Alignment(horizontal='center')
        elif col in [5]:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')

widths_analise = [8, 40, 20, 14, 18, 12, 14, 12, 12]
for i, width in enumerate(widths_analise, 1):
    ws_analise.column_dimensions[get_column_letter(i)].width = width

# Salvar arquivo
filename = f'c:/Users/Fiscal/PROJETOS/mcp.local/Relatorio_TLLF_COMPLETO_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(filename)

print(f"\n{'='*60}")
print(f"✓ RELATÓRIO COMPLETO GERADO COM SUCESSO!")
print(f"{'='*60}")
print(f"\n📄 Arquivo: {filename}")
print(f"\n📊 Estatísticas:")
print(f"   • Total de registros: {len(df)}")
print(f"   • Contribuintes únicos: {df['id_contribuinte'].nunique()}")
print(f"   • Registros 2025: {len(stats_2025)}")
print(f"   • Registros 2024: {len(stats_2024)}")
print(f"   • Registros em aberto: {len(df_aberto)}")
print(f"\n📋 Sheets criadas:")
print(f"   1. Resumo Geral")
print(f"   2. TLLF 2025 Completo ({len(stats_2025)} registros)")
print(f"   3. TLLF 2024 Completo ({len(stats_2024)} registros)")
print(f"   4. Em Aberto ({len(df_aberto)} registros)")
print(f"   5. Análise por Contribuinte ({len(df_agrupado)} contribuintes)")
print(f"\n✅ Status corretamente mapeados:")
print(f"   • Status 1 = ABERTO (Amarelo)")
print(f"   • Status 2 = CANCELADO (Cinza)")
print(f"   • Status 5 = PAGO (Verde)")
print(f"   • Status 13 = PARCELADO (Azul Claro)")
print(f"{'='*60}\n")
