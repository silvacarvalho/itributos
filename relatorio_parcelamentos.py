 import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("=" * 80)
print("RELATÓRIO DE PARCELAMENTOS 2022-2025")
print("=" * 80)

# Conectar ao banco
conn = psycopg2.connect(
    host="localhost",
    port=5432,
    database="itributos",
    user="postgres",
    password="postgres"
)

print("✓ Conectado ao banco de dados")

# Query principal - Buscar todos os parcelamentos
query = """
WITH parcelamentos AS (
    SELECT 
        ao.id as parcelamento_id,
        ao.person_id,
        p.name as contribuinte_nome,
        p.cpf_cnpj,
        EXTRACT(YEAR FROM ao.date_agreement) as ano_parcelamento,
        ao.date_agreement,
        ao.parcel_number as qtd_parcelas,
        ao.initial_due_date as primeira_parcela,
        COALESCE(ao.tribute_value, 0) + COALESCE(ao.correction_value, 0) + 
        COALESCE(ao.interest_value, 0) + COALESCE(ao.fine_value, 0) as valor_parcelado,
        ao.downpayment_amount as valor_entrada,
        ao.class_agreement as tipo,
        a.status as status_acordo,
        a.protocol_number,
        a.created_at as data_criacao_acordo,
        ao.process_number,
        ao.comment as observacao
    FROM agreement_operations ao
    LEFT JOIN agreements a ON a.agreement_operation_id = ao.id
    LEFT JOIN unico_people p ON p.id = ao.person_id
    WHERE EXTRACT(YEAR FROM ao.date_agreement) BETWEEN 2022 AND 2025
),
receitas_parcelamento AS (
    SELECT 
        ao.id as parcelamento_id,
        STRING_AGG(DISTINCT r.acronym, ', ' ORDER BY r.acronym) as receitas_sigla,
        STRING_AGG(DISTINCT r.name, ' | ' ORDER BY r.name) as receitas_nome,
        COUNT(DISTINCT odao.payment_parcel_id) as qtd_debitos_incluidos
    FROM agreement_operations ao
    LEFT JOIN other_debts_agreement_operations odao ON odao.agreement_operation_id = ao.id
    LEFT JOIN payment_parcels pp ON pp.id = odao.payment_parcel_id
    LEFT JOIN payments pm ON pm.id = pp.payment_id
    LEFT JOIN payment_taxables pt ON pt.payment_id = pm.id
    LEFT JOIN revenues r ON r.id = pt.revenue_id
    WHERE EXTRACT(YEAR FROM ao.date_agreement) BETWEEN 2022 AND 2025
    GROUP BY ao.id
),
parcelas_acordo AS (
    SELECT 
        pm.payable_id as agreement_id,
        COUNT(*) as total_parcelas_geradas,
        COUNT(CASE WHEN pp.status = 5 THEN 1 END) as parcelas_pagas,
        COUNT(CASE WHEN pp.status = 1 THEN 1 END) as parcelas_abertas,
        COUNT(CASE WHEN pp.status = 2 THEN 1 END) as parcelas_canceladas,
        SUM(CASE WHEN pp.status = 5 THEN pp.total ELSE 0 END) as valor_pago,
        SUM(CASE WHEN pp.status = 1 THEN pp.total ELSE 0 END) as valor_aberto,
        MIN(pp.due_date) as primeira_parcela_vencimento,
        MAX(pp.due_date) as ultima_parcela_vencimento,
        MAX(pp.updated_at) as data_ultimo_pagamento
    FROM payments pm
    JOIN payment_parcels pp ON pp.payment_id = pm.id
    WHERE pm.payable_type = 'Agreement'
    GROUP BY pm.payable_id
),
historico_reparcelamento AS (
    SELECT 
        person_id,
        COUNT(DISTINCT parcelamento_id) as qtd_parcelamentos,
        STRING_AGG(DISTINCT protocol_number::text, ', ' ORDER BY protocol_number::text) as protocolos,
        MIN(date_agreement) as primeiro_parcelamento,
        MAX(date_agreement) as ultimo_parcelamento
    FROM parcelamentos
    GROUP BY person_id
    HAVING COUNT(DISTINCT parcelamento_id) > 1
)
SELECT 
    par.*,
    rp.receitas_sigla,
    rp.receitas_nome,
    rp.qtd_debitos_incluidos,
    COALESCE(pa.total_parcelas_geradas, 0) as parcelas_geradas,
    COALESCE(pa.parcelas_pagas, 0) as parcelas_pagas,
    COALESCE(pa.parcelas_abertas, 0) as parcelas_abertas,
    COALESCE(pa.parcelas_canceladas, 0) as parcelas_canceladas,
    COALESCE(pa.valor_pago, 0) as valor_pago,
    COALESCE(pa.valor_aberto, 0) as valor_aberto,
    pa.primeira_parcela_vencimento,
    pa.ultima_parcela_vencimento,
    pa.data_ultimo_pagamento,
    hr.qtd_parcelamentos as total_parcelamentos_contribuinte,
    hr.protocolos as todos_protocolos,
    hr.primeiro_parcelamento as data_primeiro_acordo,
    hr.ultimo_parcelamento as data_ultimo_acordo,
    CASE 
        WHEN hr.qtd_parcelamentos > 1 THEN 'Sim'
        ELSE 'Não'
    END as possui_reparcelamento,
    CASE
        WHEN par.status_acordo = 'canceled' THEN 'Cancelado'
        WHEN par.status_acordo = 'paid' THEN 'Quitado'
        WHEN par.status_acordo = 'opened' AND pa.parcelas_pagas > 0 THEN 'Em dia'
        WHEN par.status_acordo = 'opened' AND pa.parcelas_abertas > 0 THEN 'Em aberto'
        ELSE 'Indefinido'
    END as situacao_atual
FROM parcelamentos par
LEFT JOIN receitas_parcelamento rp ON rp.parcelamento_id = par.parcelamento_id
LEFT JOIN parcelas_acordo pa ON pa.agreement_id = (
    SELECT id FROM agreements WHERE agreement_operation_id = par.parcelamento_id LIMIT 1
)
LEFT JOIN historico_reparcelamento hr ON hr.person_id = par.person_id
ORDER BY par.ano_parcelamento DESC, par.date_agreement DESC
"""

print("✓ Executando consulta...")
df = pd.read_sql_query(query, conn)
conn.close()

print(f"✓ Dados carregados: {len(df)} parcelamentos")

# Converter datas
date_columns = ['date_agreement', 'primeira_parcela', 'data_criacao_acordo', 
                'primeira_parcela_vencimento', 'ultima_parcela_vencimento', 
                'data_ultimo_pagamento', 'data_primeiro_acordo', 'data_ultimo_acordo']

for col in date_columns:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors='coerce')

# Criar Workbook
wb = Workbook()
wb.remove(wb.active)

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
quitado_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
emdia_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
aberto_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
cancelado_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
reparcelamento_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============== SHEET 1: RESUMO GERAL ==============
print("✓ Criando sheet: Resumo Geral")
ws_resumo = wb.create_sheet("Resumo Geral")

# Título
ws_resumo.merge_cells('A1:F1')
ws_resumo['A1'] = 'RELATÓRIO DE PARCELAMENTOS 2022-2025'
ws_resumo['A1'].font = Font(size=16, bold=True, color="366092")
ws_resumo['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_resumo.row_dimensions[1].height = 30

# Estatísticas por ano
ws_resumo['A3'] = 'Resumo por Ano'
ws_resumo['A3'].font = Font(size=14, bold=True)

headers_ano = ['Ano', 'Qtd Parcelamentos', 'Contribuintes Únicos', 'Valor Total', 
               'Quitados', 'Em Andamento', 'Cancelados']
for col, header in enumerate(headers_ano, 1):
    cell = ws_resumo.cell(row=5, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row_idx = 6
for ano in sorted(df['ano_parcelamento'].unique(), reverse=True):
    dados_ano = df[df['ano_parcelamento'] == ano]
    
    ws_resumo.cell(row=row_idx, column=1, value=int(ano))
    ws_resumo.cell(row=row_idx, column=2, value=len(dados_ano))
    ws_resumo.cell(row=row_idx, column=3, value=dados_ano['person_id'].nunique())
    
    valor_total = dados_ano['valor_parcelado'].sum()
    ws_resumo.cell(row=row_idx, column=4, 
                   value=f"R$ {valor_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    
    quitados = len(dados_ano[dados_ano['status_acordo'] == 'paid'])
    ws_resumo.cell(row=row_idx, column=5, value=quitados)
    
    andamento = len(dados_ano[dados_ano['status_acordo'] == 'opened'])
    ws_resumo.cell(row=row_idx, column=6, value=andamento)
    
    cancelados = len(dados_ano[dados_ano['status_acordo'] == 'canceled'])
    ws_resumo.cell(row=row_idx, column=7, value=cancelados)
    
    for col in range(1, 8):
        ws_resumo.cell(row=row_idx, column=col).border = border
        ws_resumo.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center')
    
    row_idx += 1

# Ajustar larguras
widths_resumo = [10, 20, 20, 20, 15, 18, 15]
for i, width in enumerate(widths_resumo, 1):
    ws_resumo.column_dimensions[get_column_letter(i)].width = width

# Reparcelamentos
ws_resumo[f'A{row_idx+2}'] = 'Análise de Reparcelamentos'
ws_resumo[f'A{row_idx+2}'].font = Font(size=14, bold=True)

reparcelados = df[df['possui_reparcelamento'] == 'Sim']
ws_resumo[f'A{row_idx+4}'] = f'Total de contribuintes com reparcelamento:'
ws_resumo[f'B{row_idx+4}'] = reparcelados['person_id'].nunique()
ws_resumo[f'B{row_idx+4}'].font = Font(bold=True, size=12, color="C00000")

ws_resumo[f'A{row_idx+5}'] = f'Total de parcelamentos refeitos:'
ws_resumo[f'B{row_idx+5}'] = len(reparcelados)
ws_resumo[f'B{row_idx+5}'].font = Font(bold=True, size=12, color="C00000")

# ============== SHEET 2: TODOS OS PARCELAMENTOS ==============
print("✓ Criando sheet: Todos os Parcelamentos")
ws_todos = wb.create_sheet("Todos os Parcelamentos")

headers_todos = ['ID', 'Protocolo', 'Contribuinte', 'CPF/CNPJ', 'Receitas', 'Débitos', 
                 'Ano', 'Data Acordo', 'Qtd Parcelas', 'Valor Parcelado', 'Status', 'Situação', 
                 'Parcelas Pagas', 'Parcelas Abertas', 'Valor Pago', 'Valor Aberto',
                 'Possui Reparc.']

for col, header in enumerate(headers_todos, 1):
    cell = ws_todos.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

row_idx = 2
for _, row_data in df.iterrows():
    dados = [
        row_data['parcelamento_id'],
        row_data['protocol_number'] if pd.notna(row_data['protocol_number']) else '-',
        row_data['contribuinte_nome'],
        row_data['cpf_cnpj'],
        row_data['receitas_sigla'] if pd.notna(row_data['receitas_sigla']) else '-',
        int(row_data['qtd_debitos_incluidos']) if pd.notna(row_data['qtd_debitos_incluidos']) else 0,
        int(row_data['ano_parcelamento']),
        row_data['date_agreement'].strftime('%d/%m/%Y') if pd.notna(row_data['date_agreement']) else '-',
        int(row_data['qtd_parcelas']),
        f"R$ {row_data['valor_parcelado']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        row_data['status_acordo'] if pd.notna(row_data['status_acordo']) else 'Sem status',
        row_data['situacao_atual'],
        int(row_data['parcelas_pagas']),
        int(row_data['parcelas_abertas']),
        f"R$ {row_data['valor_pago']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        f"R$ {row_data['valor_aberto']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        row_data['possui_reparcelamento']
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_todos.cell(row=row_idx, column=col, value=value)
        cell.border = border
        
        # Aplicar cores
        if row_data['situacao_atual'] == 'Quitado':
            cell.fill = quitado_fill
        elif row_data['situacao_atual'] == 'Em dia':
            cell.fill = emdia_fill
        elif row_data['situacao_atual'] == 'Cancelado':
            cell.fill = cancelado_fill
        elif row_data['possui_reparcelamento'] == 'Sim':
            cell.fill = reparcelamento_fill
        elif row_data['situacao_atual'] == 'Em aberto':
            cell.fill = aberto_fill
        
        if col in [1, 6, 7, 9, 13, 14]:
            cell.alignment = Alignment(horizontal='center')
        elif col in [10, 15, 16]:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')
    
    row_idx += 1

# Ajustar larguras
widths_todos = [8, 10, 30, 18, 20, 10, 8, 12, 12, 18, 12, 15, 12, 14, 16, 16, 15]
for i, width in enumerate(widths_todos, 1):
    ws_todos.column_dimensions[get_column_letter(i)].width = width

# ============== SHEET 3: REPARCELAMENTOS ==============
print("✓ Criando sheet: Reparcelamentos")
ws_reparc = wb.create_sheet("Reparcelamentos")

df_reparc = df[df['possui_reparcelamento'] == 'Sim'].copy()

headers_reparc = ['Contribuinte', 'CPF/CNPJ', 'Receitas Parceladas', 'Qtd Total Parcelamentos', 
                  'Todos os Protocolos', 'Primeiro Acordo', 'Último Acordo',
                  'Protocolo Atual', 'Status Atual', 'Situação']

for col, header in enumerate(headers_reparc, 1):
    cell = ws_reparc.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

row_idx = 2
for _, row_data in df_reparc.iterrows():
    dados = [
        row_data['contribuinte_nome'],
        row_data['cpf_cnpj'],
        row_data['receitas_sigla'] if pd.notna(row_data['receitas_sigla']) else '-',
        int(row_data['total_parcelamentos_contribuinte']),
        row_data['todos_protocolos'] if pd.notna(row_data['todos_protocolos']) else '-',
        row_data['data_primeiro_acordo'].strftime('%d/%m/%Y') if pd.notna(row_data['data_primeiro_acordo']) else '-',
        row_data['data_ultimo_acordo'].strftime('%d/%m/%Y') if pd.notna(row_data['data_ultimo_acordo']) else '-',
        row_data['protocol_number'] if pd.notna(row_data['protocol_number']) else '-',
        row_data['status_acordo'] if pd.notna(row_data['status_acordo']) else 'Sem status',
        row_data['situacao_atual']
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_reparc.cell(row=row_idx, column=col, value=value)
        cell.border = border
        cell.fill = reparcelamento_fill
        
        if col in [4, 8]:
            cell.alignment = Alignment(horizontal='center')
        elif col in [6, 7]:
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(horizontal='left')
    
    row_idx += 1

# Ajustar larguras
widths_reparc = [30, 18, 25, 22, 25, 15, 15, 15, 15, 15]
for i, width in enumerate(widths_reparc, 1):
    ws_reparc.column_dimensions[get_column_letter(i)].width = width

# ============== SHEET 4: ANÁLISE POR CONTRIBUINTE ==============
print("✓ Criando sheet: Análise por Contribuinte")
ws_contrib = wb.create_sheet("Análise por Contribuinte")

df_contrib = df.groupby(['person_id', 'contribuinte_nome', 'cpf_cnpj']).agg({
    'parcelamento_id': 'count',
    'valor_parcelado': 'sum',
    'valor_pago': 'sum',
    'valor_aberto': 'sum',
    'parcelas_pagas': 'sum',
    'parcelas_abertas': 'sum',
    'possui_reparcelamento': lambda x: 'Sim' if 'Sim' in x.values else 'Não'
}).reset_index()

df_contrib.columns = ['ID', 'Contribuinte', 'CPF/CNPJ', 'Qtd Parcelamentos', 
                      'Valor Total Parcelado', 'Valor Total Pago', 'Valor Total Aberto',
                      'Total Parcelas Pagas', 'Total Parcelas Abertas', 'Possui Reparcelamento']

headers_contrib = df_contrib.columns.tolist()

for col, header in enumerate(headers_contrib, 1):
    cell = ws_contrib.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = border
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

row_idx = 2
for _, row_data in df_contrib.iterrows():
    dados = [
        int(row_data['ID']),
        row_data['Contribuinte'],
        row_data['CPF/CNPJ'],
        int(row_data['Qtd Parcelamentos']),
        f"R$ {row_data['Valor Total Parcelado']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        f"R$ {row_data['Valor Total Pago']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        f"R$ {row_data['Valor Total Aberto']:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
        int(row_data['Total Parcelas Pagas']),
        int(row_data['Total Parcelas Abertas']),
        row_data['Possui Reparcelamento']
    ]
    
    for col, value in enumerate(dados, 1):
        cell = ws_contrib.cell(row=row_idx, column=col, value=value)
        cell.border = border
        
        if row_data['Possui Reparcelamento'] == 'Sim':
            cell.fill = reparcelamento_fill
        
        if col in [1, 4, 8, 9, 10]:
            cell.alignment = Alignment(horizontal='center')
        elif col in [5, 6, 7]:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left')
    
    row_idx += 1

# Ajustar larguras
widths_contrib = [8, 30, 18, 18, 20, 18, 18, 18, 20, 20]
for i, width in enumerate(widths_contrib, 1):
    ws_contrib.column_dimensions[get_column_letter(i)].width = width

# Salvar arquivo
filename = f'c:/Users/Fiscal/PROJETOS/mcp.local/Relatorio_Parcelamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(filename)

print(f"\n{'='*80}")
print(f"✓ RELATÓRIO GERADO COM SUCESSO!")
print(f"{'='*80}")
print(f"\n📄 Arquivo: {filename}")
print(f"\n📊 Estatísticas:")
print(f"   • Total de parcelamentos: {len(df)}")
print(f"   • Contribuintes únicos: {df['person_id'].nunique()}")
print(f"   • Parcelamentos 2022: {len(df[df['ano_parcelamento'] == 2022])}")
print(f"   • Parcelamentos 2023: {len(df[df['ano_parcelamento'] == 2023])}")
print(f"   • Parcelamentos 2024: {len(df[df['ano_parcelamento'] == 2024])}")
print(f"   • Parcelamentos 2025: {len(df[df['ano_parcelamento'] == 2025])}")
print(f"   • Contribuintes com reparcelamento: {reparcelados['person_id'].nunique()}")
print(f"   • Total de reparcelamentos: {len(reparcelados)}")
print(f"\n📋 Sheets criadas:")
print(f"   1. Resumo Geral")
print(f"   2. Todos os Parcelamentos ({len(df)} registros)")
print(f"   3. Reparcelamentos ({len(df_reparc)} registros)")
print(f"   4. Análise por Contribuinte ({len(df_contrib)} contribuintes)")
print(f"\n🎨 Código de cores:")
print(f"   • Verde = Quitado")
print(f"   • Azul = Em dia")
print(f"   • Amarelo = Em aberto")
print(f"   • Cinza = Cancelado")
print(f"   • Vermelho = Com reparcelamento")
print(f"{'='*80}\n")
